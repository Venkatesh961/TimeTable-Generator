import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)
LECTURE_DURATION = 3  # 1.5 hours = 3 slots (30 mins each)
LAB_DURATION = 4      # 2 hours = 4 slots (30 mins each)

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        # Skip break times (10:30-11:00 and 12:30-14:30)
        is_morning_break = time(10, 30) <= current < time(11, 0)
        is_lunch_break = time(12, 30) <= current < time(14, 30)
        if not is_morning_break and not is_lunch_break:
            next_time = current_time + timedelta(minutes=30)
            slots.append((current, next_time.time()))
        current_time += timedelta(minutes=30)
    return slots

# Load data from Excel
try:
    df = pd.read_excel('combined.xlsx', sheet_name='Sheet1')
except FileNotFoundError:
    print("Error: File 'combined.xlsx' not found in the current directory")
    exit()

def generate_all_timetables():
    TIME_SLOTS = generate_time_slots()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    professor_schedule = {}   # Track professor assignments
    classroom_schedule = {}   # Track classroom assignments
    
    for department in df['Department'].unique():
        for semester in df[df['Department'] == department]['Semester'].unique():
            courses = df[(df['Department'] == department) & (df['Semester'] == semester)].copy()
            
            if courses.empty:
                continue
            
            # Create worksheet for this department-semester
            ws = wb.create_sheet(title=f"{department}_{semester}")
            
            # Initialize timetable structure
            timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                         for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
            
            # Process labs first (since they have more constraints)
            lab_courses = courses[courses['P'] > 0]
            for _, course in lab_courses.iterrows():
                code = str(course['Course Code'])
                name = str(course['Course Name'])
                faculty = str(course['Faculty'])
                classroom = str(course['Classroom'])
                p = int(course['P'])
                
                if faculty not in professor_schedule:
                    professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                if classroom not in classroom_schedule:
                    classroom_schedule[classroom] = {day: set() for day in range(len(DAYS))}
                
                # Schedule labs (2 hours)
                for _ in range(p):
                    scheduled = False
                    attempts = 0
                    while not scheduled and attempts < 1000:
                        day = random.randint(0, len(DAYS)-1)
                        if len(TIME_SLOTS) >= LAB_DURATION:
                            start_slot = random.randint(0, len(TIME_SLOTS)-LAB_DURATION)
                            
                            # Check if all required slots are free
                            slots_free = True
                            for i in range(LAB_DURATION):
                                if (start_slot+i in professor_schedule[faculty][day] or 
                                    start_slot+i in classroom_schedule[classroom][day] or
                                    timetable[day][start_slot+i]['type'] is not None):
                                    slots_free = False
                                    break
                            
                            if slots_free:
                                # Mark professor and classroom as busy
                                for i in range(LAB_DURATION):
                                    professor_schedule[faculty][day].add(start_slot+i)
                                    classroom_schedule[classroom][day].add(start_slot+i)
                                    timetable[day][start_slot+i]['type'] = 'LAB'
                                    timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                    timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                    timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                    timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                scheduled = True
                        attempts += 1
            
            # Process lectures and tutorials
            other_courses = courses[courses['P'] == 0]
            for _, course in other_courses.iterrows():
                code = str(course['Course Code'])
                name = str(course['Course Name'])
                faculty = str(course['Faculty'])
                classroom = str(course['Classroom'])
                l = int(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                
                if faculty not in professor_schedule:
                    professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                if classroom not in classroom_schedule:
                    classroom_schedule[classroom] = {day: set() for day in range(len(DAYS))}
                
                # Schedule lectures (1.5 hours)
                for _ in range(l):
                    scheduled = False
                    attempts = 0
                    while not scheduled and attempts < 1000:
                        day = random.randint(0, len(DAYS)-1)
                        if len(TIME_SLOTS) >= LECTURE_DURATION:
                            start_slot = random.randint(0, len(TIME_SLOTS)-LECTURE_DURATION)
                            
                            # Check if all required slots are free
                            slots_free = True
                            for i in range(LECTURE_DURATION):
                                if (start_slot+i in professor_schedule[faculty][day] or 
                                    start_slot+i in classroom_schedule[classroom][day] or
                                    timetable[day][start_slot+i]['type'] is not None):
                                    slots_free = False
                                    break
                            
                            if slots_free:
                                # Mark professor and classroom as busy
                                for i in range(LECTURE_DURATION):
                                    professor_schedule[faculty][day].add(start_slot+i)
                                    classroom_schedule[classroom][day].add(start_slot+i)
                                    timetable[day][start_slot+i]['type'] = 'LEC'
                                    timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                    timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                    timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                    timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                scheduled = True
                        attempts += 1
                
                # Schedule tutorials (30 mins)
                for _ in range(t):
                    scheduled = False
                    attempts = 0
                    while not scheduled and attempts < 1000:
                        day = random.randint(0, len(DAYS)-1)
                        slot = random.randint(0, len(TIME_SLOTS)-1)
                        
                        if (slot not in professor_schedule[faculty][day] and 
                            slot not in classroom_schedule[classroom][day] and
                            timetable[day][slot]['type'] is None):
                            
                            professor_schedule[faculty][day].add(slot)
                            classroom_schedule[classroom][day].add(slot)
                            timetable[day][slot]['type'] = 'TUT'
                            timetable[day][slot]['code'] = code
                            timetable[day][slot]['name'] = name
                            timetable[day][slot]['faculty'] = faculty
                            timetable[day][slot]['classroom'] = classroom
                            scheduled = True
                        attempts += 1
            
            # Write timetable to worksheet with merged cells
            # Create header
            header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
            ws.append(header)
            
            # Apply header formatting
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Fill data and merge cells
            lec_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            lab_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            tut_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                          top=Side(style='thin'), bottom=Side(style='thin'))
            
            for day_idx, day in enumerate(DAYS):
                row_num = day_idx + 2  # +1 for header, +1 because rows start at 1
                ws.append([day])
                
                # Track merged regions
                merge_ranges = []
                current_merge = None
                
                for slot_idx in range(len(TIME_SLOTS)):
                    cell_value = ''
                    cell_fill = None
                    
                    if timetable[day_idx][slot_idx]['type']:
                        if timetable[day_idx][slot_idx]['code']:  # First slot of activity
                            activity_type = timetable[day_idx][slot_idx]['type']
                            if activity_type == 'LEC':
                                duration = LECTURE_DURATION
                                cell_fill = lec_fill
                            elif activity_type == 'LAB':
                                duration = LAB_DURATION
                                cell_fill = lab_fill
                            else:  # TUT
                                duration = 1
                                cell_fill = tut_fill
                            
                            # Create merged range
                            start_col = get_column_letter(slot_idx + 2)  # +1 for day column
                            end_col = get_column_letter(slot_idx + duration + 1)
                            merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                            merge_ranges.append(merge_range)
                            
                            # Prepare cell value
                            code = timetable[day_idx][slot_idx]['code']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            cell_value = f"{code} {activity_type}\n{classroom}"
                
                    # Write to cell
                    cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                    if cell_fill:
                        cell.fill = cell_fill
                    cell.border = border
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                
                # Apply all merges for this row
                for merge_range in merge_ranges:
                    ws.merge_cells(merge_range)
            
            # Adjust column widths and row heights
            for col_idx in range(1, len(TIME_SLOTS)+2):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15
            
            for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                ws.row_dimensions[row[0].row].height = 40
    
    # Save the workbook
    wb.save("all_timetables_merged.xlsx")
    print("All timetables with merged cells saved to all_timetables_merged.xlsx")

if __name__ == "__main__":
    generate_all_timetables()