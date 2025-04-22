from flask import Flask, render_template, request, send_file, flash, redirect, url_for, session, jsonify
import os
from werkzeug.utils import secure_filename
import timetable_generator_0 as generator
import pandas as pd
import csv
from zipfile import ZipFile
import time
import shutil
import glob
import faculty_timetable as ft
from openpyxl import load_workbook
import json
from datetime import datetime
from werkzeug.serving import WSGIRequestHandler

# Update server timeout
WSGIRequestHandler.protocol_version = "HTTP/1.1"

app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Required for flash messages
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['UPLOADED_TIMETABLE'] = None  # Add new session key for tracking uploaded timetable
app.config['TIMEOUT'] = 300  # 5 minutes timeout

# Ensure upload folder exists
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

@app.route('/')
def index():
    courses_uploaded = os.path.exists('tt data/combined.csv')
    courses = []
    departments = []
    semesters = []
    
    if courses_uploaded:
        try:
            df = pd.read_csv('tt data/combined.csv', encoding='utf-8-sig')
            courses = df.to_dict('records')
            departments = sorted(df['Department'].unique())
            
            # Custom sorting for semesters 
            semesters = sorted(df['Semester'].unique())
        except Exception as e:
            print(f"Error loading data: {str(e)}")
            courses_uploaded = False
    
    # Return JSON if requested
    if request.args.get('fetch_courses'):
        return jsonify({
            'courses': courses,
            'departments': departments,
            'semesters': semesters
        })
    
    return render_template('index.html', 
                         courses_uploaded=courses_uploaded,
                         courses=courses,
                         departments=departments,
                         semesters=semesters)

@app.route('/view-courses')
def view_courses():
    try:
        df = pd.read_csv('tt data/combined.csv', encoding='utf-8-sig')
        courses = df.to_dict('records')
        departments = sorted(df['Department'].unique())
        return render_template('courses.html', courses=courses, departments=departments)
    except:
        flash('No courses data available. Please upload a file first.')
        return redirect(url_for('index'))

@app.route('/unscheduled')
def unscheduled_courses():
    unscheduled = session.get('unscheduled', [])
    return render_template('unscheduled.html', courses=unscheduled)

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return {'success': False, 'error': 'No file uploaded'}
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.csv'):
        return {'success': False, 'error': 'Invalid file - must be CSV'}
    
    try:
        # Ensure directory exists
        os.makedirs('tt data', exist_ok=True)
        file.save('tt data/combined.csv')
        
        # Validate data structure with updated required columns
        df = pd.read_csv('tt data/combined.csv', encoding='utf-8-sig')
        required_columns = ['Department', 'Semester', 'Course Code', 'Course Name', 
                          'L', 'T', 'P', 'S', 'C', 'Faculty']
        if not all(col in df.columns for col in required_columns):
            return {'success': False, 'error': 'Invalid course file format'}
        
        # Clean semester data to remove section information
        df['Semester'] = df['Semester'].astype(str).str.extract('(\d+)').astype(int)
        df.to_csv('tt data/combined.csv', index=False)
            
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/upload-rooms', methods=['POST'])
def upload_rooms():
    if 'file' not in request.files:
        return {'success': False, 'error': 'No file uploaded'}
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.csv'):
        return {'success': False, 'error': 'Invalid file'}
    
    try:
        file.save('rooms.csv')
        # Validate room data
        with open('rooms.csv', 'r') as f:
            reader = csv.DictReader(f)
            rooms = list(reader)
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/upload-batches', methods=['POST'])
def upload_batches():
    if 'file' not in request.files:
        return {'success': False, 'error': 'No file uploaded'}
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.csv'):
        return {'success': False, 'error': 'Invalid file'}
    
    try:
        # Ensure directory exists
        os.makedirs('tt data', exist_ok=True)
        file.save('tt data/updated_batches.csv')
        
        # Validate batch data
        df = pd.read_csv('tt data/updated_batches.csv')
        required_columns = ['Department', 'Semester', 'Total_Students', 'MaxBatchSize']
        if not all(col in df.columns for col in required_columns):
            return {'success': False, 'error': 'Invalid batch file format'}
            
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/upload-reserved', methods=['POST'])
def upload_reserved():
    if 'file' not in request.files:
        return {'success': False, 'error': 'No file uploaded'}
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.csv'):
        return {'success': False, 'error': 'Invalid file'}
    
    try:
        # Ensure directory exists
        os.makedirs('tt data', exist_ok=True)
        file.save('tt data/reserved_slots.csv')
        
        # Validate reserved slots data
        df = pd.read_csv('tt data/reserved_slots.csv')
        required_columns = ['Day', 'Start Time', 'End Time', 'Semester']
        if not all(col in df.columns for col in required_columns):
            return {'success': False, 'error': 'Invalid reserved slots file format'}
            
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/upload-faculty', methods=['POST'])
def upload_faculty():
    if 'file' not in request.files:
        return {'success': False, 'error': 'No file uploaded'}
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.csv'):
        return {'success': False, 'error': 'Invalid file'}
    
    try:
        # Ensure directory exists
        os.makedirs('tt data', exist_ok=True)
        file.save('tt data/FACULTY.csv')
        
        # Validate faculty data structure
        df = pd.read_csv('tt data/FACULTY.csv')
        required_columns = ['Faculty ID', 'Name', 'Preferred Days', 'Preferred Times']
        if not all(col in df.columns for col in required_columns):
            return {'success': False, 'error': 'Invalid faculty file format'}
            
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/upload-elective-registrations', methods=['POST'])
def upload_elective_registrations():
    if 'file' not in request.files:
        return {'success': False, 'error': 'No file uploaded'}
    
    file = request.files['file']
    if file.filename == '' or not file.filename.endswith('.csv'):
        return {'success': False, 'error': 'Invalid file'}
    
    try:
        # Ensure directory exists
        os.makedirs('tt data', exist_ok=True)
        file.save('tt data/elective_registrations.csv')
        
        # Validate elective registration data
        df = pd.read_csv('tt data/elective_registrations.csv')
        required_columns = ['Course Code', 'Total Students']
        if not all(col in df.columns for col in required_columns):
            return {'success': False, 'error': 'Invalid elective registrations file format'}
            
        return {'success': True}
    except Exception as e:
        return {'success': False, 'error': str(e)}

@app.route('/generate', methods=['POST'])
def generate():
    # Check all required files exist
    required_files = [
        ('tt data/combined.csv', 'Course data'),
        ('tt data/FACULTY.csv', 'Faculty data'),
        ('tt data/elective_registrations.csv', 'Elective registrations'),
        ('rooms.csv', 'Room data'),
        ('tt data/updated_batches.csv', 'Batch data')
    ]
    
    missing_files = []
    for file_path, file_name in required_files:
        if not os.path.exists(file_path):
            missing_files.append(file_name)
    
    if missing_files:
        flash(f'Missing required files: {", ".join(missing_files)}')
        return redirect(url_for('index'))
    
    try:
        # Generate timetables and get list of generated files
        timetable_files = generator.generate_all_timetables()
        
        if not timetable_files:
            flash('No timetables were generated')
            return redirect(url_for('index'))

        # Create a unique zip filename using timestamp
        timestamp = int(time.time())
        zip_filename = f'timetables_{timestamp}.zip'
        
        # Create zip file containing all timetables
        with ZipFile(zip_filename, 'w') as zipf:
            for file in timetable_files:
                if os.path.exists(file):
                    zipf.write(file)
                    try:
                        os.remove(file)
                    except:
                        pass

        return send_file(
            zip_filename,
            as_attachment=True,
            download_name='department_timetables.zip',
            mimetype='application/zip'
        )
        
    except Exception as e:
        flash(f'Error generating timetables: {str(e)}')
        return redirect(url_for('index'))

@app.route('/faculty-view')
def faculty_view():
    upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'timetables')
    faculty_list = set()
    timetables_uploaded = False

    # Check if directory exists and has Excel files
    if os.path.exists(upload_dir) and glob.glob(os.path.join(upload_dir, '*.xlsx')):
        timetables_uploaded = True
        try:
            for file in glob.glob(os.path.join(upload_dir, '*.xlsx')):
                try:
                    wb = load_workbook(file, read_only=True)
                    for sheet in wb.worksheets:
                        # Skip sheets that don't look like timetables
                        if not sheet['A2'].value in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']:
                            continue
                        
                        # Scan cells for faculty names
                        for row in range(2, 7):  # 5 days
                            for col in range(2, sheet.max_column + 1):
                                cell = sheet.cell(row=row, column=col).value
                                if cell:
                                    # Faculty name is typically the last line
                                    lines = str(cell).strip().split('\n')
                                    if len(lines) >= 3:
                                        faculty = lines[2].strip()
                                        # Clean faculty name to remove any course code prefixes
                                        faculty = faculty.split('/')[-1].strip()  # Take last name if multiple
                                        if '(' in faculty:  # Remove anything in parentheses
                                            faculty = faculty.split('(')[0].strip()
                                        if faculty and faculty != '-':
                                            faculty_list.add(faculty)
                finally:
                    wb.close()  # Ensure workbook is closed
        except Exception as e:
            print(f"Error reading timetables: {str(e)}")

    return render_template('faculty_view.html', 
                         faculty_list=sorted(faculty_list), 
                         timetables_uploaded=timetables_uploaded)

@app.route('/upload-dept-timetables', methods=['POST'])
def upload_dept_timetables():
    if 'files[]' not in request.files:
        return jsonify({'success': False, 'error': 'No files uploaded'})
        
    files = request.files.getlist('files[]')
    upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'timetables')
    
    try:
        # Clear previous uploads with retry mechanism
        if os.path.exists(upload_dir):
            retries = 3
            for _ in range(retries):
                try:
                    shutil.rmtree(upload_dir)
                    break
                except PermissionError:
                    time.sleep(1)  # Wait before retry
            else:
                return jsonify({'success': False, 'error': 'Could not remove existing files. Please close any open Excel files.'})
                
        os.makedirs(upload_dir)
        
        # Save new files
        for file in files:
            if file.filename.endswith('.xlsx'):
                filepath = os.path.join(upload_dir, secure_filename(file.filename))
                file.save(filepath)
                # Verify file is not locked
                try:
                    with open(filepath, 'r+b') as f:
                        pass
                except PermissionError:
                    return jsonify({'success': False, 'error': f'File {file.filename} is being used by another process'})
                    
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/generate-faculty-timetable/<faculty_name>')
def generate_faculty_timetable(faculty_name):
    try:
        # Use uploaded timetables directory
        upload_dir = 'uploads/timetables'
        if not os.path.exists(upload_dir):
            flash('No timetables uploaded yet')
            return redirect(url_for('faculty_view'))

        timetable_files = glob.glob(os.path.join(upload_dir, '*.xlsx'))
        if not timetable_files:
            flash('No timetable files found')
            return redirect(url_for('faculty_view'))

        faculty_name = faculty_name.strip()
        if not faculty_name:
            flash('No faculty selected')
            return redirect(url_for('faculty_view'))
            
        wb = ft.generate_faculty_timetable(faculty_name, timetable_files)
        if not wb:
            flash('Failed to generate timetable')
            return redirect(url_for('faculty_view'))
            
        # Save to temp file with sanitized name
        safe_name = "".join(c for c in faculty_name if c.isalnum() or c in (' ', '.', '_')).rstrip()
        safe_name = safe_name.replace(' ', '_')
        filename = f'faculty_timetable_{safe_name}.xlsx'
        wb.save(filename)
        
        return send_file(
            filename,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error generating timetable: {str(e)}')
        return redirect(url_for('faculty_view'))

@app.route('/download_analytics')
def download_analytics():
    try:
        # Get list of department timetable files
        timetable_files = glob.glob('timetable_*.xlsx')
        if not timetable_files:
            flash('No timetable files found. Please generate timetables first.')
            return redirect(url_for('faculty_view'))
            
        # Generate analytics report
        from analytics import generate_analytics_report
        analytics_wb = generate_analytics_report(timetable_files)
        
        if not analytics_wb:
            flash('Error generating analytics report')
            return redirect(url_for('faculty_view'))
            
        # Save workbook to memory
        from io import BytesIO
        output = BytesIO()
        analytics_wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='timetable_analytics.xlsx'
        )
        
    except Exception as e:
        flash(f'Error generating analytics: {str(e)}')
        return redirect(url_for('faculty_view'))

@app.route('/save-config', methods=['POST'])
def save_config():
    try:
        config = request.get_json()
        with open('config.json', 'w') as f:
            json.dump(config, f, indent=4)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# Add cleanup on startup
def cleanup_old_files():
    try:
        if os.path.exists('cleanup.txt'):
            with open('cleanup.txt', 'r') as f:
                files = f.readlines()
            
            # Try to remove each file
            remaining_files = []
            for file in files:
                file = file.strip()
                try:
                    if os.path.exists(file):
                        os.remove(file)
                except:
                    remaining_files.append(file)
            
            # Update cleanup.txt with files that couldn't be removed
            if remaining_files:
                with open('cleanup.txt', 'w') as f:
                    f.writelines(remaining_files)
            else:
                os.remove('cleanup.txt')
    except:
        pass

if __name__ == '__main__':
    cleanup_old_files()  # Run cleanup on startup
    app.run(debug=True)
