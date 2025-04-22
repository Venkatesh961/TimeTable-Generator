import pandas as pd
from datetime import datetime, time
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

def get_time_slot(slot_num):
    """Convert slot number to time range string"""
    # Start from 9:00 AM, each slot is 30 minutes
    start_hour = 9 + (slot_num - 1) // 2
    start_min = "30" if (slot_num - 1) % 2 else "00"
    end_hour = 9 + slot_num // 2
    end_min = "00" if slot_num % 2 else "30"
    
    return f"{start_hour:02d}:{start_min}-{end_hour:02d}:{end_min}"

def parse_cell_contents(cell_value):
    """Parse contents from timetable cell"""
    if not cell_value or 'BREAK' in str(cell_value):
        return None
        
    parts = str(cell_value).strip().split('\n')
    if len(parts) >= 3:
        course_parts = parts[0].split()
        return {
            'course_code': course_parts[0] if course_parts else '',
            'type': course_parts[1] if len(course_parts) > 1 else '',
            'room': parts[1].strip(),
            'faculty': parts[2].strip()
        }
    return None

def extract_timetable_data(sheet, dept):
    """Extract timetable data from worksheet"""
    if sheet['A2'].value not in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']:
        return None
        
    slots_data = []
    for row in range(2, 7):  # Monday to Friday
        day = sheet.cell(row=row, column=1).value
        day_slots = []
        
        # Process each timeslot
        for col in range(2, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col).value
            if cell:
                # Parse cell content (Course\nRoom\nFaculty)
                parts = str(cell).strip().split('\n')
                if len(parts) >= 3:
                    day_slots.append({
                        'course': parts[0].strip(),
                        'room': parts[1].strip(),
                        'faculty': parts[2].strip(),
                        'slot': col-1
                    })
        slots_data.append({
            'day': day,
            'slots': day_slots
        })
    
    return slots_data

def generate_room_usage_report(timetable_files, rooms_df):
    """Generate room usage analytics"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Room Usage Analysis"
    
    headers = ['Room', 'Type', 'Capacity', 'Total Hours Used', 'Utilization %', 
              'Peak Usage Day', 'Peak Hours', 'Free Time Slots']
    ws.append(headers)

    # Track usage by day and time
    room_usage = {}  # {room_id: {day: set(time_slots)}}
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    for file in timetable_files:
        try:
            # Read Excel file
            df = pd.read_excel(file, header=0)
            
            # Validate DataFrame structure
            if df.empty or df.shape[0] < 7 or df.shape[1] < 20:
                print(f"Skipping {file} - Invalid format")
                continue
            
            # Get time slots from header row
            time_slots = df.columns[1:].tolist()
            
            # Process days (rows 1-5 contain the actual schedule)
            for row in range(0, 5):  # Process Monday to Friday
                day = str(df.iloc[row, 0])  # Get day name from first column
                if not day in days:
                    continue
                    
                # Process each time slot
                for col in range(1, df.shape[1]):  # Skip first column (day names)
                    try:
                        cell_data = parse_cell_contents(df.iloc[row, col])
                        if cell_data and cell_data['room']:
                            room = cell_data['room']
                            if room not in room_usage:
                                room_usage[room] = {d: set() for d in days}
                            room_usage[room][day].add(col)
                    except IndexError as ie:
                        print(f"Index error in file {file} at row {row}, col {col}: {str(ie)}")
                        continue
                    except Exception as e:
                        print(f"Error processing cell in file {file} at row {row}, col {col}: {str(e)}")
                        continue
                        
        except Exception as e:
            print(f"Error processing file {file}: {str(e)}")
            continue

    # Generate statistics
    total_slots = 19 * 5  # 19 slots per day * 5 days
    
    for _, room in rooms_df.iterrows():
        room_id = room['id']
        if room_id in room_usage:
            # Calculate total used slots
            total_used = sum(len(slots) for slots in room_usage[room_id].values())
            
            # Find peak usage day
            peak_day = max(days, key=lambda d: len(room_usage[room_id][d]))
            peak_slots = len(room_usage[room_id][peak_day])
            
            # Get free time slots
            free_slots = []
            for day in days:
                used = sorted(room_usage[room_id][day])
                if not used:
                    free_slots.append(f"{day}: All day")
                else:
                    gaps = []
                    prev = 0
                    for slot in used:
                        if slot - prev > 1:
                            gaps.append(f"{get_time_slot(prev+1)}-{get_time_slot(slot-1)}")
                        prev = slot
                    if prev < 19:
                        gaps.append(f"{get_time_slot(prev+1)}-18:30")
                    if gaps:
                        free_slots.append(f"{day}: {', '.join(gaps)}")
            
            utilization = (total_used / total_slots) * 100
            
            ws.append([
                room_id,
                room['type'],
                room['capacity'],
                f"{total_used/2:.1f} hours",
                f"{utilization:.1f}%", 
                peak_day,
                f"{peak_slots/2:.1f} hours",
                "\n".join(free_slots)
            ])
        else:
            # Room never used
            ws.append([
                room_id,
                room['type'], 
                room['capacity'],
                "0 hours",
                "0%",
                "N/A",
                "0 hours",
                "All slots free"
            ])

    return wb

def generate_faculty_schedule_report(timetable_files, faculty_df):
    """Generate faculty schedule analytics"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Faculty Analysis"
    
    headers = ['Faculty', 'Total Teaching Hours', 'Classes Per Day', 
              'Peak Teaching Day', 'Course Distribution', 'Free Time Slots']
    ws.append(headers)

    faculty_schedule = {}  # {faculty: {day: {slot: course_info}}}
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']

    for file in timetable_files:
        try:
            df = pd.read_excel(file, header=0)
            sheet_name = df.iloc[0,0].split('_')  # Get department and semester from sheet name
            dept = sheet_name[0] if len(sheet_name) > 0 else ''
            
            for row in range(1, 6):  # Skip header row, process 5 days
                day = str(df.iloc[row, 0])
                if day not in days:
                    continue
                    
                for col in range(1, df.shape[1]):
                    cell_data = parse_cell_contents(df.iloc[row, col])
                    if cell_data:
                        faculty = cell_data['faculty']
                        if faculty:
                            if faculty not in faculty_schedule:
                                faculty_schedule[faculty] = {d: {} for d in days}
                            
                            faculty_schedule[faculty][day][col] = {
                                'course': cell_data['course_code'],
                                'type': cell_data['type'],
                                'dept': dept
                            }

        except Exception as e:
            print(f"Error processing {file}: {str(e)}")
            continue

    # Generate statistics for each faculty
    for _, faculty in faculty_df.iterrows():
        name = faculty['Name']
        if name in faculty_schedule:
            schedule = faculty_schedule[name]
            
            # Calculate workload
            total_hours = 0
            course_types = {'LEC': 0, 'LAB': 0, 'TUT': 0}
            courses_handled = {}  # {course: department}
            
            for day, slots in schedule.items():
                for _, info in slots.items():
                    if info['type'] == 'LEC':
                        total_hours += 1.5  # 1.5 hours for lecture
                    elif info['type'] == 'LAB':
                        total_hours += 2  # 2 hours for lab
                    elif info['type'] == 'TUT':
                        total_hours += 1  # 1 hour for tutorial
                        
                    course_types[info['type']] += 1
                    courses_handled[info['course']] = info['dept']
            
            # Calculate peak day
            peak_day = max(days, key=lambda d: len(schedule[d]))
            peak_slots = len(schedule[peak_day])
            
            # Calculate average classes per day
            active_days = sum(1 for d in days if schedule[d])
            avg_classes = len([s for d in days for s in schedule[d]])/active_days if active_days else 0
            
            # Format course distribution
            course_dist = [
                f"{len(courses_handled)} courses ({', '.join(f'{c}({d})' for c,d in courses_handled.items())})",
                f"LEC: {course_types['LEC']}, LAB: {course_types['LAB']}, TUT: {course_types['TUT']}"
            ]
            
            # Get free time slots
            free_slots = []
            for day in days:
                used_slots = sorted(schedule[day].keys())
                if not used_slots:
                    free_slots.append(f"{day}: All day")
                else:
                    gaps = []
                    prev = 0
                    for slot in used_slots:
                        if slot - prev > 1:
                            gaps.append(f"{get_time_slot(prev+1)}-{get_time_slot(slot-1)}")
                        prev = slot
                    if prev < 19:
                        gaps.append(f"{get_time_slot(prev+1)}-18:30")
                    if gaps:
                        free_slots.append(f"{day}: {', '.join(gaps)}")
            
            ws.append([
                name,
                f"{total_hours:.1f} hours",
                f"{avg_classes:.1f}",
                f"{peak_day} ({peak_slots} classes)",
                "\n".join(course_dist),
                "\n".join(free_slots)
            ])
        else:
            ws.append([name, "0 hours", "0", "N/A", "No courses", "All slots free"])

    return wb

def generate_analytics_report(timetable_files):
    """Generate combined analytics report"""
    try:
        rooms_df = pd.read_csv('rooms.csv')
        faculty_df = pd.read_csv('tt data/FACULTY.csv')
    except Exception as e:
        print(f"Error loading data files: {e}")
        return None

    # Generate reports
    wb = generate_room_usage_report(timetable_files, rooms_df)
    faculty_wb = generate_faculty_schedule_report(timetable_files, faculty_df)
    
    # Copy faculty sheet to main workbook
    faculty_sheet = faculty_wb.active
    ws = wb.create_sheet("Faculty Analysis")
    
    # Copy headers and data
    for row in faculty_sheet.iter_rows():
        ws.append([cell.value for cell in row])
    
    # Format workbook
    for sheet in wb.worksheets:
        # Format headers
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            
        # Adjust column widths
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            sheet.column_dimensions[col[0].column_letter].width = max_length + 2

    return wb
