import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import glob
import os

# Add color palette for courses
COLORS = [
    "FFB6C1", "98FB98", "87CEFA", "DDA0DD", "F0E68C", 
    "E6E6FA", "FFDAB9", "B0E0E6", "FFA07A", "D8BFD8",
    "AFEEEE", "F08080", "90EE90", "ADD8E6", "FFB6C1"
]

def generate_faculty_timetable(faculty_name, timetable_files):
    """Generate a consolidated timetable for a specific faculty"""
    wb = Workbook()
    ws = wb.active
    ws.title = faculty_name.replace('/', '_').replace('\\', '_')[:31]
    
    # Track time slots and days
    time_slots = None
    days = []
    course_colors = {}  # Map courses to colors
    color_idx = 0
    
    # Process first file to get structure
    if timetable_files:
        first_wb = load_workbook(timetable_files[0], read_only=True)
        first_sheet = first_wb.worksheets[0]
        
        header = [cell.value for cell in first_sheet[1][1:]]
        time_slots = header
        
        for row in range(2, 7):
            day = first_sheet.cell(row=row, column=1).value
            if day:
                days.append(day)
    
    if not time_slots or not days:
        return None
        
    header_row = ['Day'] + time_slots
    ws.append(header_row)
    
    # Initialize schedule with None to track merging
    schedule = {day: [{'content': '', 'code': None, 'duration': 0} for _ in time_slots] for day in days}
    
    # Process all timetable files
    for file in timetable_files:
        dept_wb = load_workbook(file, read_only=True)
        
        for sheet in dept_wb.worksheets:
            for row in range(2, len(days) + 2):
                day = sheet.cell(row=row, column=1).value
                if not day in days:
                    continue
                
                for col in range(2, len(time_slots) + 2):
                    cell = sheet.cell(row=row, column=col)
                    if cell.value and faculty_name in str(cell.value):
                        details = str(cell.value).split('\n')
                        if len(details) >= 2:
                            code_type = details[0]
                            room = details[1]
                            course_code = code_type.split()[0]
                            
                            # Extract semester and section from sheet name
                            sheet_info = sheet.title.split('_')
                            dept = sheet_info[0]
                            sem = sheet_info[1]
                            section = sheet_info[2] if len(sheet_info) > 2 else ''
                            sem_sec = f"{dept} {sem}"
                            if section:
                                sem_sec += f"-{section}"
                            
                            # Determine duration from activity type
                            duration = 1
                            if 'LEC' in code_type:
                                duration = 3  # 1.5 hours
                            elif 'LAB' in code_type:
                                duration = 4  # 2 hours
                            elif 'TUT' in code_type:
                                duration = 2  # 1 hour
                            
                            # Assign color if new course
                            if course_code not in course_colors:
                                course_colors[course_code] = COLORS[color_idx % len(COLORS)]
                                color_idx += 1
                            
                            # Store in schedule with duration and semester info
                            col_idx = col - 2
                            if col_idx + duration <= len(time_slots):
                                schedule[day][col_idx] = {
                                    'content': f"{code_type}\n{room}\n{sem_sec}",  # Add semester info
                                    'code': course_code,
                                    'duration': duration
                                }
                                # Mark subsequent slots as taken
                                for i in range(1, duration):
                                    schedule[day][col_idx + i] = {
                                        'content': '',
                                        'code': course_code,
                                        'duration': 0
                                    }
    
    # Write schedule to worksheet with merging
    for day_idx, day in enumerate(days, 2):
        ws.cell(row=day_idx, column=1, value=day)
        col_idx = 2
        while col_idx <= len(time_slots) + 1:
            slot_info = schedule[day][col_idx - 2]
            if slot_info['duration'] > 0:
                cell = ws.cell(row=day_idx, column=col_idx, value=slot_info['content'])
                if slot_info['duration'] > 1:
                    # Merge cells
                    end_col = get_column_letter(col_idx + slot_info['duration'] - 1)
                    ws.merge_cells(f"{get_column_letter(col_idx)}{day_idx}:{end_col}{day_idx}")
                
                # Apply color
                if slot_info['code']:
                    cell.fill = PatternFill(start_color=course_colors[slot_info['code']], 
                                          end_color=course_colors[slot_info['code']], 
                                          fill_type="solid")
                col_idx += slot_info['duration']
            else:
                col_idx += 1

    # Apply formatting
    header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Format header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Format cells
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value).split('\n')[0]))
            except:
                pass
        adjusted_width = max(12, min(max_length + 2, 20))
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Set row heights
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 60  # Increased height for 3 lines
        
    return wb
