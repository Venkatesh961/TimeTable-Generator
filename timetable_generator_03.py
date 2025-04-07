import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)
LECTURE_DURATION = 3  # 1.5 hours = 3 slots (30 mins each)
LAB_DURATION = 4      # 2 hours = 4 slots (30 mins each)

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        # Skip break times (10:30-11:00 and 12:30-14:30)
        is_morning_break = time(10, 30) <= current < time(11, 0)
        is_lunch_break = time(12, 30) <= current < time(14, 30)
        if not is_morning_break and not is_lunch_break:
            next_time = current_time + timedelta(minutes=30)
            slots.append((current, next_time.time()))
        current_time += timedelta(minutes=30)
    return slots

# Load data from Excel
try:
    df = pd.read_excel('combined.xlsx', sheet_name='Sheet1')
except FileNotFoundError:
    print("Error: File 'combined.xlsx' not found in the current directory")
    exit()

def generate_all_timetables():
    TIME_SLOTS = generate_time_slots()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    professor_schedule = {}   # Track professor assignments
    classroom_schedule = {}   # Track classroom assignments
    
    for department in df['Department'].unique():
        for semester in df[df['Department'] == department]['Semester'].unique():
            courses = df[(df['Department'] == department) & (df['Semester'] == semester)].copy()
            
            if courses.empty:
                continue
            
            # Create worksheet for this department-semester
            ws = wb.create_sheet(title=f"{department}_{semester}")
            
            # Initialize timetable structure
            timetable = {day: {slot: [] for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
            
            # Process labs first (since they have more constraints)
            lab_courses = courses[courses['P'] > 0]
            for _, course in lab_courses.iterrows():
                code = str(course['Course Code'])
                name = str(course['Course Name'])
                faculty = str(course['Faculty'])
                classroom = str(course['Classroom'])
                p = int(course['P'])
                
                if faculty not in professor_schedule:
                    professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                if classroom not in classroom_schedule:
                    classroom_schedule[classroom] = {day: set() for day in range(len(DAYS))}
                
                # Schedule labs (2 hours)
                for _ in range(p):
                    scheduled = False
                    attempts = 0
                    while not scheduled and attempts < 1000:
                        day = random.randint(0, len(DAYS)-1)
                        # Find a starting slot that allows for LAB_DURATION consecutive slots
                        if len(TIME_SLOTS) >= LAB_DURATION:
                            slot = random.randint(0, len(TIME_SLOTS)-LAB_DURATION)
                            
                            # Check if all required slots are free
                            slots_free = True
                            for i in range(LAB_DURATION):
                                if (slot+i in professor_schedule[faculty][day] or 
                                    slot+i in classroom_schedule[classroom][day] or
                                    any(c[0] == code for c in timetable[day][slot+i])):
                                    slots_free = False
                                    break
                            
                            if slots_free:
                                # Mark professor and classroom as busy
                                for i in range(LAB_DURATION):
                                    professor_schedule[faculty][day].add(slot+i)
                                    classroom_schedule[classroom][day].add(slot+i)
                                    if i == 0:
                                        timetable[day][slot+i].append((code, name, 'LAB', faculty, classroom))
                                    else:
                                        timetable[day][slot+i].append(('', '', 'cont.', '', ''))
                                scheduled = True
                        attempts += 1
            
            # Process lectures and tutorials
            other_courses = courses[courses['P'] == 0]
            for _, course in other_courses.iterrows():
                code = str(course['Course Code'])
                name = str(course['Course Name'])
                faculty = str(course['Faculty'])
                classroom = str(course['Classroom'])
                l = int(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                
                if faculty not in professor_schedule:
                    professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                if classroom not in classroom_schedule:
                    classroom_schedule[classroom] = {day: set() for day in range(len(DAYS))}
                
                # Schedule lectures (1.5 hours)
                for _ in range(l):
                    scheduled = False
                    attempts = 0
                    while not scheduled and attempts < 1000:
                        day = random.randint(0, len(DAYS)-1)
                        if len(TIME_SLOTS) >= LECTURE_DURATION:
                            slot = random.randint(0, len(TIME_SLOTS)-LECTURE_DURATION)
                            
                            # Check if all required slots are free
                            slots_free = True
                            for i in range(LECTURE_DURATION):
                                if (slot+i in professor_schedule[faculty][day] or 
                                    slot+i in classroom_schedule[classroom][day] or
                                    any(c[0] == code for c in timetable[day][slot+i])):
                                    slots_free = False
                                    break
                            
                            if slots_free:
                                # Mark professor and classroom as busy
                                for i in range(LECTURE_DURATION):
                                    professor_schedule[faculty][day].add(slot+i)
                                    classroom_schedule[classroom][day].add(slot+i)
                                    if i == 0:
                                        timetable[day][slot+i].append((code, name, 'LEC', faculty, classroom))
                                    else:
                                        timetable[day][slot+i].append(('', '', 'cont.', '', ''))
                                scheduled = True
                        attempts += 1
                
                # Schedule tutorials (30 mins)
                for _ in range(t):
                    scheduled = False
                    attempts = 0
                    while not scheduled and attempts < 1000:
                        day = random.randint(0, len(DAYS)-1)
                        slot = random.randint(0, len(TIME_SLOTS)-1)
                        
                        if (slot not in professor_schedule[faculty][day] and 
                            slot not in classroom_schedule[classroom][day] and
                            not any(c[0] == code for c in timetable[day][slot])):
                            
                            professor_schedule[faculty][day].add(slot)
                            classroom_schedule[classroom][day].add(slot)
                            timetable[day][slot].append((code, name, 'TUT', faculty, classroom))
                            scheduled = True
                        attempts += 1
            
            # Write timetable to worksheet
            # Create header
            header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
            ws.append(header)
            
            # Fill data
            for day_idx, day in enumerate(DAYS):
                row = [day]
                for slot_idx in range(len(TIME_SLOTS)):
                    if timetable[day_idx][slot_idx]:
                        course = timetable[day_idx][slot_idx][0]
                        if course[0]:  # Not continuation
                            display = f"{course[0]} {course[2]}\n{course[4]}"
                        else:
                            display = ""
                        row.append(display)
                    else:
                        row.append("")
                ws.append(row)
            
            # Apply formatting
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            lec_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            lab_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
            tut_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    if cell.row == 1:  # Header row
                        cell.fill = header_fill
                        cell.font = Font(bold=True)
                    elif cell.value and isinstance(cell.value, str):
                        if 'LEC' in cell.value:
                            cell.fill = lec_fill
                        elif 'LAB' in cell.value:
                            cell.fill = lab_fill
                        elif 'TUT' in cell.value:
                            cell.fill = tut_fill
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = min(adjusted_width, 30)
    
    # Save the workbook
    wb.save("all_timetables.xlsx")
    print("All timetables saved to all_timetables.xlsx")

if __name__ == "__main__":
    generate_all_timetables()