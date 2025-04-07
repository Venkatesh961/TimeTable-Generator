import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import csv
import glob
import os

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)

# Duration constants
HOUR_SLOTS = 2  # Number of 30-min slots that make 1 hour
LECTURE_DURATION = 3  # 1.5 hours = 3 slots (30 mins each) 
LAB_DURATION = 4      # 2 hours = 4 slots (30 mins each)
TUTORIAL_DURATION = 2 # 1 hour = 2 slots
SELF_STUDY_DURATION = 2  # 1 hour = 2 slots
BREAK_DURATION = 1    # 30 mins = 1 slot

# Generate time slots once at module level
TIME_SLOTS = []

def initialize_time_slots():
    global TIME_SLOTS
    TIME_SLOTS = generate_time_slots()

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        next_time = current_time + timedelta(minutes=30)
        
        # Keep all time slots but we'll mark break times later
        slots.append((current, next_time.time()))
        current_time = next_time
    
    return slots

def load_rooms():
    rooms = {}
    try:
        with open('rooms.csv', 'r') as f:
            reader = csv.DictReader(f)
            for row in reader:
                rooms[row['id']] = {
                    'capacity': int(row['capacity']),
                    'type': row['type'],
                    'roomNumber': row['roomNumber'],
                    'schedule': {day: set() for day in range(len(DAYS))}
                }
    except FileNotFoundError:
        print("Warning: rooms.csv not found, using default room allocation")
        return None
    return rooms

def load_batch_data():
    """Load batch information and calculate sections automatically"""
    batch_info = {}
    try:
        df = pd.read_csv('tt data/updated_batches.csv')
        for _, row in df.iterrows():
            total_students = row['Total_Students']
            max_batch_size = row['MaxBatchSize']
            
            # Calculate number of sections needed
            num_sections = (total_students + max_batch_size - 1) // max_batch_size
            section_size = (total_students + num_sections - 1) // num_sections

            batch_info[(row['Department'], row['Semester'])] = {
                'total': total_students,
                'num_sections': num_sections,
                'section_size': section_size
            }
    except FileNotFoundError:
        print("Warning: updated_batches.csv not found, using default batch sizes")
        return None
    return batch_info

def find_adjacent_lab_room(room_id, rooms):
    """Find an adjacent lab room based on room numbering"""
    if not room_id:
        return None
    
    # Get room number and extract base info
    current_num = int(''.join(filter(str.isdigit, rooms[room_id]['roomNumber'])))
    current_floor = current_num // 100
    
    # Look for adjacent room with same type
    for rid, room in rooms.items():
        if rid != room_id and room['type'] == rooms[room_id]['type']:
            room_num = int(''.join(filter(str.isdigit, room['roomNumber'])))
            # Check if on same floor and adjacent number
            if room_num // 100 == current_floor and abs(room_num - current_num) == 1:
                return rid
    return None

def find_suitable_room(course_type, department, semester, day, start_slot, duration, rooms, batch_info, course_code="", used_rooms=None):
    """Find suitable room(s) considering batch sizes and avoiding room conflicts"""
    if not rooms:
        return "DEFAULT_ROOM"
    
    required_capacity = 60  # Default fallback
    if batch_info:
        dept_info = batch_info.get((department, semester))
        if dept_info:
            if course_type in ['COMPUTER_LAB', 'HARDWARE_LAB']:
                required_capacity = dept_info['section_size']
            else:
                required_capacity = dept_info['section_size']

    # Track currently used rooms for this time slot if provided
    used_room_ids = set() if used_rooms is None else used_rooms
    
    # First try to find a single room that fits everyone and isn't being used
    for room_id, room in rooms.items():
        # Skip if room is already in use by another section
        if room_id in used_room_ids:
            continue

        # Skip library rooms entirely
        if room['type'].upper() == 'LIBRARY':
            continue
            
        # Match specific lab types based on course code and room type
        if course_type == 'COMPUTER_LAB':
            if room['type'].upper() != 'COMPUTER_LAB':
                continue
        elif course_type == 'HARDWARE_LAB':
            if room['type'].upper() != 'HARDWARE_LAB':
                continue
        elif course_type in ['LEC', 'TUT', 'SS']:
            if not ('LECTURE_ROOM' in room['type'].upper() or 'SEATER' in room['type'].upper()):
                continue
                
        # Check capacity
        if room['capacity'] >= required_capacity:
            # Check availability for all slots needed
            slots_free = True
            for i in range(duration):
                if start_slot + i in room['schedule'][day]:
                    slots_free = False
                    break
                    
            if slots_free:
                # Reserve the room
                for i in range(duration):
                    room['schedule'][day].add(start_slot + i)
                return room_id

    # If no single room fits and it's a lab, try to find adjacent rooms
    if course_type in ['COMPUTER_LAB', 'HARDWARE_LAB']:
        available_rooms = []
        for room_id, room in rooms.items():
            if room_id not in used_room_ids and room['type'].upper() == course_type:
                slots_free = True
                for i in range(duration):
                    if start_slot + i in room['schedule'][day]:
                        slots_free = False
                        break
                if slots_free:
                    available_rooms.append(room_id)
        
        # Try single large room first
        for room_id in available_rooms:
            if rooms[room_id]['capacity'] >= required_capacity:
                for i in range(duration):
                    rooms[room_id]['schedule'][day].add(start_slot + i)
                return room_id
        
        # Try adjacent room pairs
        for room_id in available_rooms:
            adjacent_room = find_adjacent_lab_room(room_id, rooms)
            if adjacent_room and adjacent_room in available_rooms:
                combined_capacity = rooms[room_id]['capacity'] + rooms[adjacent_room]['capacity']
                if combined_capacity >= required_capacity:
                    for i in range(duration):
                        rooms[room_id]['schedule'][day].add(start_slot + i)
                        rooms[adjacent_room]['schedule'][day].add(start_slot + i)
                    return f"{room_id},{adjacent_room}"

    return None

def get_required_room_type(course):
    """Determine required room type based on course attributes"""
    if pd.notna(course['P']) and course['P'] > 0:
        course_code = str(course['Course Code']).upper()
        # For CS courses, use computer labs
        if 'CS' in course_code or 'DS' in course_code:
            return 'COMPUTER_LAB'
        # For EC courses, use hardware labs
        elif 'EC' in course_code:
            return 'HARDWARE_LAB'
        return 'COMPUTER_LAB'  # Default to computer lab if unspecified
    else:
        # For lectures, tutorials, and self-study
        return 'LECTURE_ROOM'

# Load data from CSV with robust error handling
try:
    # Try different encodings and handle BOM
    encodings_to_try = ['utf-8-sig', 'utf-8', 'cp1252']
    df = None
    last_error = None
    
    for encoding in encodings_to_try:
        try:
            df = pd.read_csv('tt data/combined.csv', encoding=encoding)
            # Convert empty strings and 'nan' strings to actual NaN
            df = df.replace(r'^\s*$', pd.NA, regex=True)
            df = df.replace('nan', pd.NA)
            break
        except UnicodeDecodeError:
            continue
        except Exception as e:
            last_error = e
            continue
            
    if df is None:
        print(f"Error: Unable to read combined.csv. Please check the file format.\nDetails: {str(last_error)}")
        exit()
        
except Exception as e:
    print(f"Error: Failed to load combined.csv.\nDetails: {str(e)}")
    exit()

if df.empty:
    print("Error: No data found in combined.csv")
    exit()

def is_break_time(slot):
    """Check if a time slot falls within break times"""
    start, end = slot
    # Morning break: 10:30-11:00
    morning_break = (time(10, 30) <= start < time(11, 0))
    # Lunch break: 12:30-14:30
    lunch_break = (time(12, 30) <= start < time(14, 30))
    return morning_break or lunch_break

def is_lecture_scheduled(timetable, day, start_slot, end_slot):
    """Check if there's a lecture scheduled in the given time range"""
    for slot in range(start_slot, end_slot):
        if (slot < len(timetable[day]) and 
            timetable[day][slot]['type'] and 
            timetable[day][slot]['type'] in ['LEC', 'LAB', 'TUT']):
            return True
    return False

def calculate_required_slots(course):
    """Calculate how many slots needed based on L, T, P, S values and credits"""
    l = float(course['L']) if pd.notna(course['L']) else 0  # Lecture credits
    t = int(course['T']) if pd.notna(course['T']) else 0    # Tutorial hours
    p = int(course['P']) if pd.notna(course['P']) else 0    # Lab hours
    s = int(course['S']) if pd.notna(course['S']) else 0    # Self study hours
    c = int(course['C']) if pd.notna(course['C']) else 0    # Total credits
    
    # Check if course is self-study only
    if s > 0 and l == 0 and t == 0 and p == 0:
        return 0, 0, 0, 0
        
    # Calculate number of lecture sessions based on credits
    lecture_sessions = 0
    if l > 0:
        # For 3 credits = 2 sessions of 1.5 hours each
        # For 2 credits = 1 session of 1.5 hours plus a 1 hour session
        # For 1 credit = 1 session of 1.5 hours
        lecture_sessions = max(1, round(l * 2/3))  # Scale credits to sessions
    
    # Other calculations remain the same
    tutorial_sessions = t  
    lab_sessions = p // 2  # 2 hours per lab session
    self_study_sessions = s // 4 if (l > 0 or t > 0 or p > 0) else 0
    
    return lecture_sessions, tutorial_sessions, lab_sessions, self_study_sessions

def select_faculty(faculty_str):
    """Select a faculty from potentially multiple options."""
    if '/' in faculty_str:
        # Split by slash and strip whitespace
        faculty_options = [f.strip() for f in faculty_str.split('/')]
        return faculty_options[0]  # Take first faculty as default
    return faculty_str

def check_faculty_daily_components(professor_schedule, faculty, day, department, semester, section, timetable, course_code=None, activity_type=None):
    """Check faculty/course scheduling constraints for the day"""
    component_count = 0
    lecture_courses = set()  # Track courses that have lectures scheduled
    
    # Check all slots for this day
    for slot in timetable[day].values():
        if slot['faculty'] == faculty and slot['type'] in ['LEC', 'LAB', 'TUT']:
            slot_code = slot.get('code', '')
            if slot_code:
                # Count components only for same semester/section
                if slot_code in df[(df['Department'] == department) & 
                                 (df['Semester'] == semester)]['Course Code'].values:
                    component_count += 1
                    # Track lecture occurrences
                    if slot['type'] == 'LEC':
                        lecture_courses.add(slot_code)

    # Check if trying to schedule another lecture of same course
    if activity_type == 'LEC' and course_code in lecture_courses:
        return False
        
    return component_count < 2  # Return True if faculty can take another component

# Add new class to track unscheduled components
class UnscheduledComponent:
    def __init__(self, department, semester, code, name, faculty, component_type, sessions, section='', reason=''):
        self.department = department
        self.semester = semester
        self.code = code
        self.name = name
        self.faculty = faculty 
        self.component_type = component_type
        self.sessions = sessions
        self.section = section
        self.reason = reason

def load_reserved_slots():
    """Load reserved time slots from CSV file"""
    try:
        reserved_slots_path = os.path.join('tt data', 'reserved_slots.csv')
        if not os.path.exists(reserved_slots_path):
            print("Warning: reserved_slots.csv not found in uploads, no slots will be reserved")
            return {day: {} for day in DAYS}
            
        df = pd.read_csv(reserved_slots_path)
        reserved = {day: {} for day in DAYS}
        
        for _, row in df.iterrows():
            day = row['Day']
            start = datetime.strptime(row['Start Time'], '%H:%M').time()
            end = datetime.strptime(row['End Time'], '%H:%M').time()
            department = str(row['Department'])
            # Handle semester sections (e.g., "4" matches "4A" and "4B")
            semesters = []
            for s in str(row['Semester']).split(';'):
                s = s.strip()
                if s.isdigit():  # If just a number like "4"
                    base_sem = int(s)  
                    semesters.extend([f"{base_sem}A", f"{base_sem}B", str(base_sem)])
                else:  # If already has section like "4A"
                    semesters.append(s)
            
            key = (department, tuple(semesters))
            if day not in reserved:
                reserved[day] = {}
            if key not in reserved[day]:
                reserved[day][key] = []
                
            reserved[day][key].append((start, end))
        return reserved
    except Exception as e:
        print(f"Warning: Error loading reserved slots: {str(e)}")
        return {day: {} for day in DAYS}

def is_slot_reserved(slot, day, semester, department, reserved_slots):
    """Check if a time slot is reserved for this semester and department"""
    if day not in reserved_slots:
        return False
        
    slot_start, slot_end = slot
    
    # Check each reservation
    for (dept, semesters), slots in reserved_slots[day].items():
        # Match if department is ALL or matches exactly
        if dept == 'ALL' or dept == department:
            # Match if semester is in the expanded semester list
            if str(semester) in semesters or any(str(semester).startswith(s) for s in semesters):
                for reserved_start, reserved_end in slots:
                    if (slot_start >= reserved_start and slot_start < reserved_end) or \
                       (slot_end > reserved_start and slot_end <= reserved_end):
                        return True
    return False

# Add these helper functions after existing helper functions
def get_course_priority(course):
    """Calculate course scheduling priority based on constraints"""
    priority = 0
    if pd.notna(course['P']) and course['P'] > 0:
        priority += 3  # Labs get highest priority
    if pd.notna(course['L']) and course['L'] > 2:
        priority += 2  # More lecture hours = higher priority
    if pd.notna(course['T']) and course['T'] > 0:
        priority += 1  # Tutorials add complexity
    return priority

def get_best_slots(timetable, professor_schedule, faculty, day, duration, reserved_slots, semester, department):
    """Find best available consecutive slots in a day"""
    best_slots = []
    for start_slot in range(len(TIME_SLOTS) - duration + 1):
        slots_free = True
        # Check each slot in the duration
        for i in range(duration):
            current_slot = start_slot + i
            if (current_slot in professor_schedule[faculty][day] or 
                timetable[day][current_slot]['type'] is not None or
                is_break_time(TIME_SLOTS[current_slot]) or
                is_slot_reserved(TIME_SLOTS[current_slot], DAYS[day], semester, department, reserved_slots)):
                slots_free = False
                break
        if slots_free:
            best_slots.append(start_slot)
    return best_slots

def generate_all_timetables():
    initialize_time_slots()  # Initialize time slots before using
    reserved_slots = load_reserved_slots()
    workbooks = {}  # Dictionary to store workbook for each department
    professor_schedule = {}   # Track professor assignments
    rooms = load_rooms()
    batch_info = load_batch_data()

    # Add tracking for unscheduled components
    unscheduled_components = []

    # Color palette for subjects (will cycle through these)
    subject_colors = [
        "FFB6C1", "98FB98", "87CEFA", "DDA0DD", "F0E68C", 
        "E6E6FA", "FFDAB9", "B0E0E6", "FFA07A", "D8BFD8",
        "AFEEEE", "F08080", "90EE90", "ADD8E6", "FFB6C1"
    ]

    # Add a list to track self-study only courses
    self_study_courses = []

    for department in df['Department'].unique():
        # Create new workbook for each department
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        workbooks[department] = wb
        
        # Track assigned faculty for courses
        course_faculty_assignments = {}
        
        # Process all semesters for this department
        for semester in df[df['Department'] == department]['Semester'].unique():
            courses = df[(df['Department'] == department) & (df['Semester'] == semester)].copy()
            
            if courses.empty:
                continue

            # Get section info
            dept_info = batch_info.get((department, semester))
            num_sections = dept_info['num_sections'] if dept_info else 1

            # First identify self-study only courses
            for _, course in courses.iterrows():
                l = int(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                p = int(course['P']) if pd.notna(course['P']) else 0
                s = int(course['S']) if pd.notna(course['S']) else 0
                
                if s > 0 and l == 0 and t == 0 and p == 0:
                    self_study_courses.append({
                        'code': str(course['Course Code']),
                        'name': str(course['Course Name']),
                        'faculty': str(course['Faculty']),
                        'department': department,
                        'semester': semester
                    })

            for section in range(num_sections):
                section_title = f"{department}_{semester}" if num_sections == 1 else f"{department}_{semester}_{chr(65+section)}"
                ws = wb.create_sheet(title=section_title)
                
                # Initialize timetable structure
                timetable = {day: {slot: {'type': None, 'code': '', 'name': '', 'faculty': '', 'classroom': ''} 
                         for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
                
                # Create a mapping for subject colors
                subject_color_map = {}
                course_faculty_map = {}  # For legend
                color_idx = 0
                
                # Assign colors to each unique subject
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    if code not in subject_color_map and code and code != 'nan':
                        subject_color_map[code] = subject_colors[color_idx % len(subject_colors)]
                        course_faculty_map[code] = {
                            'name': str(course['Course Name']),
                            'faculty': str(course['Faculty'])
                        }
                        color_idx += 1

                # Sort courses by priority
                courses['priority'] = courses.apply(get_course_priority, axis=1)
                courses = courses.sort_values('priority', ascending=False)

                # Process all courses - both lab and non-lab
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    
                    # Skip basket courses (B1, B2, etc)
                    if not any(code.startswith(f'B{i}') for i in range(1, 10)):
                        # For same course in different sections, try to use different faculty
                        if code in course_faculty_assignments:
                            # If multiple faculty available, try to pick a different one
                            if '/' in faculty:
                                faculty_options = [f.strip() for f in faculty.split('/')] 
                                # Remove already assigned faculty
                                available_faculty = [f for f in faculty_options 
                                                     if f not in course_faculty_assignments[code]]
                                if available_faculty:
                                    faculty = available_faculty[0]
                                else:
                                    faculty = select_faculty(faculty)
                        else:
                            faculty = select_faculty(faculty)
                            course_faculty_assignments[code] = [faculty]
                    else:
                        faculty = select_faculty(faculty)
                    
                    lecture_sessions, tutorial_sessions, lab_sessions, _ = calculate_required_slots(course)
                    
                    if faculty not in professor_schedule:
                        professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}

                    # Schedule lectures with tracking
                    for _ in range(lecture_sessions):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day = random.randint(0, len(DAYS)-1)
                            start_slot = random.randint(0, len(TIME_SLOTS)-LECTURE_DURATION)
                            
                            # Check if any slot in the range is reserved
                            slots_reserved = any(is_slot_reserved(TIME_SLOTS[start_slot + i], 
                                                                DAYS[day],
                                                                semester,
                                                                department,
                                                                reserved_slots) 
                                               for i in range(LECTURE_DURATION))
                            
                            if slots_reserved:
                                attempts += 1
                                continue
                            
                            # Check faculty daily component limit and lecture constraints
                            if not check_faculty_daily_components(professor_schedule, faculty, day, 
                                                               department, semester, section, timetable,
                                                               code, 'LEC'):
                                attempts += 1
                                continue
                                
                            # Check availability and ensure breaks between lectures
                            slots_free = True
                            for i in range(LECTURE_DURATION):
                                current_slot = start_slot + i
                                if (current_slot in professor_schedule[faculty][day] or 
                                    timetable[day][current_slot]['type'] is not None or
                                    is_break_time(TIME_SLOTS[current_slot])):
                                    slots_free = False
                                    break
                                
                                # Check for lectures before this slot
                                if current_slot > 0:
                                    if is_lecture_scheduled(timetable, day, 
                                                         max(0, current_slot - BREAK_DURATION), 
                                                         current_slot):
                                        slots_free = False
                                        break
                                
                                # Check for lectures after this slot
                                if current_slot < len(TIME_SLOTS) - 1:
                                    if is_lecture_scheduled(timetable, day,
                                                         current_slot + 1,
                                                         min(len(TIME_SLOTS), 
                                                             current_slot + BREAK_DURATION + 1)):
                                        slots_free = False
                                        break
                            
                            if slots_free:
                                room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                          day, start_slot, LECTURE_DURATION, 
                                                          rooms, batch_info, code)
                                
                                if room_id:
                                    classroom = room_id
                                    
                                    # Mark slots as used
                                    for i in range(LECTURE_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'LEC'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    scheduled = True
                            attempts += 1
                        if not scheduled:
                            unscheduled_components.append(
                                UnscheduledComponent(department, semester, code, name, 
                                                   faculty, 'LEC', 1, section)
                            )

                    # Schedule tutorials with tracking
                    for _ in range(tutorial_sessions):
                        scheduled = False
                        attempts = 0
                        while not scheduled and attempts < 1000:
                            day = random.randint(0, len(DAYS)-1)
                            
                            # Check faculty daily component limit for tutorials
                            if not check_faculty_daily_components(professor_schedule, faculty, day,
                                                               department, semester, section, timetable,
                                                               code, 'TUT'):
                                attempts += 1
                                continue
                                
                            start_slot = random.randint(0, len(TIME_SLOTS)-TUTORIAL_DURATION)
                            
                            # Check if any slot in the range is reserved
                            slots_reserved = any(is_slot_reserved(TIME_SLOTS[start_slot + i], 
                                                                DAYS[day],
                                                                semester,
                                                                department,
                                                                reserved_slots) 
                                               for i in range(TUTORIAL_DURATION))
                            
                            if slots_reserved:
                                attempts += 1
                                continue
                            
                            # Check availability
                            slots_free = True
                            for i in range(TUTORIAL_DURATION):
                                if (start_slot+i in professor_schedule[faculty][day] or 
                                    timetable[day][start_slot+i]['type'] is not None or
                                    is_break_time(TIME_SLOTS[start_slot+i])):
                                    slots_free = False
                                    break
                            
                            if slots_free:
                                room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                          day, start_slot, TUTORIAL_DURATION, 
                                                          rooms, batch_info, code)
                                
                                if room_id:
                                    classroom = room_id
                                    
                                    # Mark slots as used
                                    for i in range(TUTORIAL_DURATION):
                                        professor_schedule[faculty][day].add(start_slot+i)
                                        timetable[day][start_slot+i]['type'] = 'TUT'
                                        timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                        timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                        timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                        timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                    scheduled = True
                            attempts += 1
                        if not scheduled:
                            unscheduled_components.append(
                                UnscheduledComponent(department, semester, code, name,
                                                   faculty, 'TUT', 1, section)
                            )

                    # Schedule labs with tracking
                    if lab_sessions > 0:
                        room_type = get_required_room_type(course)
                        for _ in range(lab_sessions):
                            scheduled = False
                            attempts = 0
                            scheduling_reason = ""
                            
                            # Try each day in random order
                            days = list(range(len(DAYS)))
                            random.shuffle(days)
                            
                            for day in days:
                                # Get all possible slots for this day
                                possible_slots = get_best_slots(timetable, professor_schedule, 
                                                              faculty, day, LAB_DURATION, 
                                                              reserved_slots, semester, department)
                                
                                for start_slot in possible_slots:
                                    room_id = find_suitable_room(room_type, department, semester,
                                                               day, start_slot, LAB_DURATION,
                                                               rooms, batch_info, code)
                                    
                                    if room_id:
                                        classroom = room_id if ',' not in str(room_id) else f"{room_id.split(',')[0]}+{room_id.split(',')[1]}"
                                        
                                        # Mark slots as used
                                        for i in range(LAB_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'LAB'
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                        break
                                
                                if scheduled:
                                    break
                                
                            if not scheduled:
                                unscheduled_components.append(
                                    UnscheduledComponent(department, semester, code, name,
                                                       faculty, 'LAB', 1, section,
                                                       "Could not find suitable room and time slot combination")
                                )

                # Schedule self-study sessions
                for _, course in courses.iterrows():
                    code = str(course['Course Code'])
                    name = str(course['Course Name'])
                    faculty = str(course['Faculty'])
                    _, _, _, self_study_sessions = calculate_required_slots(course)
                    
                    if self_study_sessions > 0:
                        if faculty not in professor_schedule:
                            professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                        
                        # Schedule each self-study session (1 hour each)
                        for _ in range(self_study_sessions):
                            scheduled = False
                            attempts = 0
                            while not scheduled and attempts < 1000:
                                day = random.randint(0, len(DAYS)-1)
                                start_slot = random.randint(0, len(TIME_SLOTS)-SELF_STUDY_DURATION)
                                
                                # Check if any slot in the range is reserved
                                slots_reserved = any(is_slot_reserved(TIME_SLOTS[start_slot + i], 
                                                                    DAYS[day],
                                                                    semester,
                                                                    department,
                                                                    reserved_slots) 
                                                   for i in range(SELF_STUDY_DURATION))
                                
                                if slots_reserved:
                                    attempts += 1
                                    continue
                                
                                # Check availability
                                slots_free = True
                                for i in range(SELF_STUDY_DURATION):
                                    if (start_slot+i in professor_schedule[faculty][day] or 
                                        timetable[day][start_slot+i]['type'] is not None or
                                        is_break_time(TIME_SLOTS[start_slot+i])):
                                        slots_free = False
                                        break
                                
                                if slots_free:
                                    room_id = find_suitable_room('LECTURE_ROOM', department, semester, 
                                                              day, start_slot, SELF_STUDY_DURATION, 
                                                              rooms, batch_info, code)
                                    
                                    if room_id:
                                        classroom = room_id
                                        
                                        # Mark slots as used
                                        for i in range(SELF_STUDY_DURATION):
                                            professor_schedule[faculty][day].add(start_slot+i)
                                            timetable[day][start_slot+i]['type'] = 'SS'  # SS for Self Study
                                            timetable[day][start_slot+i]['code'] = code if i == 0 else ''
                                            timetable[day][start_slot+i]['name'] = name if i == 0 else ''
                                            timetable[day][start_slot+i]['faculty'] = faculty if i == 0 else ''
                                            timetable[day][start_slot+i]['classroom'] = classroom if i == 0 else ''
                                        scheduled = True
                                attempts += 1

                # Write timetable to worksheet
                header = ['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS]
                ws.append(header)
                
                header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
                header_font = Font(bold=True)
                header_alignment = Alignment(horizontal='center', vertical='center')
                
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                
                lec_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                lab_fill = PatternFill(start_color="98FB98", end_color="98FB98", fill_type="solid")
                tut_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")
                ss_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
                break_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                border = Border(left=Side(style='thin'), right=Side(style='thin'),
                              top=Side(style='thin'), bottom=Side(style='thin'))
                
                for day_idx, day in enumerate(DAYS):
                    row_num = day_idx + 2
                    ws.append([day])
                    
                    merge_ranges = []  # Track merge ranges for this row
                    
                    for slot_idx in range(len(TIME_SLOTS)):
                        cell_value = ''
                        cell_fill = None
                        
                        if is_break_time(TIME_SLOTS[slot_idx]):
                            cell_value = "BREAK"
                            cell_fill = break_fill
                        elif timetable[day_idx][slot_idx]['type']:
                            activity_type = timetable[day_idx][slot_idx]['type']
                            code = timetable[day_idx][slot_idx]['code']
                            classroom = timetable[day_idx][slot_idx]['classroom']
                            faculty = timetable[day_idx][slot_idx]['faculty']
                            
                            # Only create content for start of activity
                            if code:
                                # Get duration based on activity type
                                duration = {
                                    'LEC': LECTURE_DURATION,
                                    'LAB': LAB_DURATION,
                                    'TUT': TUTORIAL_DURATION,
                                    'SS': SELF_STUDY_DURATION
                                }.get(activity_type, 1)
                                
                                # Use subject-specific color
                                if code in subject_color_map:
                                    cell_fill = PatternFill(start_color=subject_color_map[code],
                                                          end_color=subject_color_map[code],
                                                          fill_type="solid")
                                else:
                                    cell_fill = {
                                        'LAB': lab_fill,
                                        'TUT': tut_fill,
                                        'SS': ss_fill,
                                        'LEC': lec_fill
                                    }.get(activity_type, lec_fill)
                                
                                cell_value = f"{code} {activity_type}\n{classroom}\n{faculty}"
                                
                                # Create merge range
                                if duration > 1:
                                    start_col = get_column_letter(slot_idx + 2)
                                    end_col = get_column_letter(slot_idx + duration + 1)
                                    merge_range = f"{start_col}{row_num}:{end_col}{row_num}"
                                    merge_ranges.append((merge_range, cell_fill))
                        
                        cell = ws.cell(row=row_num, column=slot_idx+2, value=cell_value)
                        if cell_fill:
                            cell.fill = cell_fill
                        cell.border = border
                        cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    
                    # Apply merges after creating all cells in the row
                    for merge_range, fill in merge_ranges:
                        ws.merge_cells(merge_range)
                        # Ensure merged cell has consistent formatting
                        merged_cell = ws[merge_range.split(':')[0]]
                        merged_cell.fill = fill
                        merged_cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

                for col_idx in range(1, len(TIME_SLOTS)+2):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 15
                
                for row in ws.iter_rows(min_row=2, max_row=len(DAYS)+1):
                    ws.row_dimensions[row[0].row].height = 40

                # Add Self-Study Only Courses section
                current_row = len(DAYS) + 4  # Initialize current_row here, before any sections

                if self_study_courses:
                    ws.cell(row=current_row, column=1, value="Self-Study Only Courses")
                    ws.cell(row=current_row, column=1).font = Font(bold=True)
                    current_row += 1
                    
                    headers = ['Course Code', 'Course Name', 'Faculty']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=current_row, column=col, value=header)
                        ws.cell(row=current_row, column=col).font = Font(bold=True)
                    current_row += 1
                    
                    for course in self_study_courses:
                        if course['department'] == department and course['semester'] == semester:
                            ws.cell(row=current_row, column=1, value=course['code'])
                            ws.cell(row=current_row, column=2, value=course['name'])
                            ws.cell(row=current_row, column=3, value=course['faculty'])
                            current_row += 1
                    
                    current_row += 2  # Add extra spacing after self-study courses

                # Add Unscheduled Components section before legend
                print(f"Total unscheduled components: {len(unscheduled_components)}")
                dept_unscheduled = [c for c in unscheduled_components 
                                    if c.department == department and 
                                    c.semester == semester and
                                    (c.section == section if num_sections > 1 else True)]
                print(f"Unscheduled for {department} semester {semester}: {len(dept_unscheduled)}")

                if unscheduled_components:
                    current_row += 4  # Add extra spacing
                    ws.cell(row=current_row, column=1, value="Unscheduled Components")
                    ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FF0000")
                    current_row += 2

                    headers = ['Course Code', 'Course Name', 'Faculty', 'Component', 'Sessions', 'Failure Reason']
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=current_row, column=col, value=header)
                        cell.font = Font(bold=True)
                        cell.border = border
                        cell.fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    current_row += 1

                    # Filter components for this department, semester and section
                    dept_unscheduled = [c for c in unscheduled_components 
                                      if c.department == department and 
                                      c.semester == semester and
                                      (c.section == section if num_sections > 1 else True)]
                    
                    for comp in dept_unscheduled:
                        cells = [
                            (comp.code, None),
                            (comp.name, None),
                            (comp.faculty, None),
                            (comp.component_type, None),
                            (comp.sessions, None),
                            (comp.reason, None)
                        ]
                        
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            if fill:
                                cell.fill = fill
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                        
                        current_row += 1

                # Improved legend formatting
                current_row += 2  # Add extra spacing before legend
                legend_title = ws.cell(row=current_row, column=1, value="Legend")
                legend_title.font = Font(bold=True, size=12)
                current_row += 2

                # Wider columns for legend
                ws.column_dimensions['A'].width = 20  # Subject Code
                ws.column_dimensions['B'].width = 40  # Subject Name
                ws.column_dimensions['C'].width = 30  # Faculty
                ws.column_dimensions['D'].width = 15  # Color

                # Add legend headers with better formatting
                legend_headers = ['Subject Code', 'Subject Name', 'Faculty', 'Color']
                for col, header in enumerate(legend_headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.border = border
                    cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1

                # Add subject entries with improved spacing
                for code, color in subject_color_map.items():
                    if code in course_faculty_map:
                        # Add spacing between rows
                        ws.row_dimensions[current_row].height = 25
                        
                        cells = [
                            (code, None),
                            (course_faculty_map[code]['name'], None),
                            (course_faculty_map[code]['faculty'], None),
                            ('', PatternFill(start_color=color, end_color=color, fill_type="solid"))
                        ]
                        
                        for col, (value, fill) in enumerate(cells, 1):
                            cell = ws.cell(row=current_row, column=col, value=value)
                            cell.border = border
                            if fill:
                                cell.fill = fill
                            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                        
                        current_row += 1

    # Save separate workbook for each department
    for department, wb in workbooks.items():
        filename = f"timetable_{department}.xlsx"
        wb.save(filename)
        print(f"Timetable for {department} saved as {filename}")

    return [f"timetable_{dept}.xlsx" for dept in workbooks.keys()]

if __name__ == "__main__":
    generate_all_timetables()