import pandas as pd
import random
from datetime import datetime, time, timedelta
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, Border, Side, Alignment

# Constants
DAYS = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
START_TIME = time(9, 0)
END_TIME = time(18, 30)
TIME_SLOT_DURATION = 30  # minutes

def generate_time_slots():
    slots = []
    current_time = datetime.combine(datetime.today(), START_TIME)
    end_time = datetime.combine(datetime.today(), END_TIME)
    
    while current_time < end_time:
        current = current_time.time()
        # Skip break times (10:30-11:00 and 12:30-14:30)
        is_morning_break = time(10, 30) <= current < time(11, 0)
        is_lunch_break = time(12, 30) <= current < time(14, 30)
        if not is_morning_break and not is_lunch_break:
            next_time = current_time + timedelta(minutes=TIME_SLOT_DURATION)
            slots.append((current, next_time.time()))
        current_time += timedelta(minutes=TIME_SLOT_DURATION)
    return slots

# Load data from Excel
try:
    df = pd.read_excel('combined.xlsx', sheet_name='Sheet1')
except FileNotFoundError:
    print("Error: File 'combined.xlsx' not found in the current directory")
    exit()

def generate_all_timetables():
    TIME_SLOTS = generate_time_slots()
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    professor_schedule = {}  # Track professor assignments to avoid conflicts
    
    for department in df['Department'].unique():
        for semester in df[df['Department'] == department]['Semester'].unique():
            courses = df[(df['Department'] == department) & (df['Semester'] == semester)].copy()
            
            if courses.empty:
                continue
            
            # Create worksheet for this department-semester
            ws = wb.create_sheet(title=f"{department}_{semester}")
            
            # Initialize timetable structure
            timetable = {day: {slot: [] for slot in range(len(TIME_SLOTS))} for day in range(len(DAYS))}
            
            for _, course in courses.iterrows():
                code = str(course['Course Code'])
                name = str(course['Course Name'])
                faculty = str(course['Faculty'])
                classroom = str(course['Classroom'])
                l = int(course['L']) if pd.notna(course['L']) else 0
                t = int(course['T']) if pd.notna(course['T']) else 0
                
                # Initialize professor in schedule if not exists
                if faculty not in professor_schedule:
                    professor_schedule[faculty] = {day: set() for day in range(len(DAYS))}
                
                # Schedule lectures (1 hour)
                for _ in range(l):
                    scheduled = False
                    attempts = 0
                    while not scheduled and attempts < 500:  # Increased attempts
                        day = random.randint(0, len(DAYS)-1)
                        slot = random.randint(0, len(TIME_SLOTS)-2)
                        
                        # Check for conflicts
                        professor_free = slot not in professor_schedule[faculty][day]
                        classroom_free = True
                        course_free = (not any(c[0] == code for c in timetable[day][slot]) and 
                                     not any(c[0] == code for c in timetable[day][slot+1]))
                        
                        if professor_free and classroom_free and course_free:
                            # Mark professor as busy
                            professor_schedule[faculty][day].add(slot)
                            professor_schedule[faculty][day].add(slot+1)
                            
                            # Add to timetable
                            timetable[day][slot].append((code, name, 'Lecture', faculty, classroom))
                            timetable[day][slot+1].append(('', '', 'cont.', '', ''))
                            scheduled = True
                        attempts += 1
                
                # Schedule tutorials (30 mins)
                for _ in range(t):
                    scheduled = False
                    attempts = 0
                    while not scheduled and attempts < 500:  # Increased attempts
                        day = random.randint(0, len(DAYS)-1)
                        slot = random.randint(0, len(TIME_SLOTS)-1)
                        
                        # Check for conflicts
                        professor_free = slot not in professor_schedule[faculty][day]
                        classroom_free = True
                        course_free = not any(c[0] == code for c in timetable[day][slot])
                        
                        if professor_free and classroom_free and course_free:
                            # Mark professor as busy
                            professor_schedule[faculty][day].add(slot)
                            
                            # Add to timetable
                            timetable[day][slot].append((code, name, 'Tutorial', faculty, classroom))
                            scheduled = True
                        attempts += 1
            
            # Write timetable to worksheet
            # Create header
            ws.append(['Day'] + [f"{slot[0].strftime('%H:%M')}-{slot[1].strftime('%H:%M')}" for slot in TIME_SLOTS])
            
            # Fill data
            for day_idx, day in enumerate(DAYS):
                row = [day]
                for slot_idx in range(len(TIME_SLOTS)):
                    if timetable[day_idx][slot_idx]:
                        course = timetable[day_idx][slot_idx][0]
                        if course[0]:  # Not continuation
                            display = f"{course[0]} {course[2][0]}\n{course[4]}"
                        else:
                            display = ""
                        row.append(display)
                    else:
                        row.append("")
                ws.append(row)
            
            # Apply formatting
            header_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                               top=Side(style='thin'), bottom=Side(style='thin'))
            
            for row in ws.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    if cell.row == 1:  # Header row
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        cell.alignment = Alignment(wrap_text=True)
            
            # Adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save("all_timetables.xlsx")
    print("All timetables saved to all_timetables.xlsx")

if __name__ == "__main__":
    generate_all_timetables()